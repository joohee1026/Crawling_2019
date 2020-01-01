import openpyxl

wb = openpyxl.Workbook()
##wb.save('test.xlsx')

##sheet = wb.active
##sheet['A1'] = 'hello world'
##sheet.cell(row=3, column=3).value = "Bye"

##subject = ['python', 'java','html','javascript']
##sheet.append(subject)

##wb.save('test2.xlsx')

sheet1 = wb.active
sheet1.title = "1st sheet"

sheet2 = wb.create_sheet("2nd sheet")

for i in range(1, 10):
    sheet1.cell(row=i,column=1).value = i
    sheet2.cell(row=1,column=i).value = i

wb.save('test3.xlsx')