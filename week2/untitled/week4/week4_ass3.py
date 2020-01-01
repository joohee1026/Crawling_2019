import requests
from bs4 import BeautifulSoup
import openpyxl

#title, author, score
li = ['mon','tue','wed','thi','fri','sat','sun']
wb = openpyxl.Workbook()
wb.create_sheet("webtoon")
wb.create_sheet("music")

sheet = wb["webtoon"]
sheet.append(['title','author','score'])
for i in li:
    raw = requests.get("https://comic.naver.com/webtoon/weekdayList.nhn?week="+str(i))
    html = BeautifulSoup(raw.text,"html.parser")
    webtoon = html.select("ul.img_list li")

    for wt in webtoon:
        title = wt.select_one("dl dt a")['title']
        author = wt.select_one("dd.desc").text
        score = wt.select_one("div.rating_type").text
        score = score[3:]
        #print(title, author, score)
        #sheet.append([title,author,score])

for i in range(1,99,50):
    print(i)
    raw = requests.get("https://www.melon.com/chart/#params%5Bidx%5D="+str(i),
                   headers={"User-Agent":"Mozilla/5.0"})
    html= BeautifulSoup(raw.text, 'html.parser')
    if i==1:
        html_parse = html.select("tbody>tr#lst50")

    elif i == 51:
        html_parse = html.select("tbody>tr#lst100")

    # >li 랑 li 의 차이! : >li는 바로 하위고 li는 쭉쭉쭉 아래까지 가능하다.
    for b in html_parse:
        title = b.select_one("div.rank01 span>a").text
        author = b.select_one("div.rank02>a").text
        album = b.select_one("div.rank03>a").text
        rank = b.select_one("div.t_center>span").text
        print("#" * 10)
        print(rank, "위")
        print("노래 제목 ", title)
        print("가수", author)
        print("앨범", album)
        print("#" * 10)

#sheet = wb["music"]
#sheet.append(['title','singer','album','like'])


##wb.save("webtoonNmusic.xlsx")
